import re
from docx import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def parse_docx(filepath):
    doc = Document(filepath)
    projects = []
    
    current_owner = None
    current_project = None
    
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
            
        # Check for owner section (e.g. "4.1 Sepehr")
        if re.match(r'^\d+\.\d+\s+([a-zA-Z]+)$', text):
            current_owner = re.match(r'^\d+\.\d+\s+([a-zA-Z]+)$', text).group(1)
            continue
            
        # Check for project title (e.g. "Sepehr - Dashboard Parameters...")
        if current_owner and text.startswith(f"{current_owner} - "):
            if current_project:
                projects.append(current_project)
            current_project = {
                'ID': f"{current_owner[:3].upper()}-{len([p for p in projects if p['Owner'] == current_owner]) + 1:02d}",
                'Item': text.replace(f"{current_owner} - ", "", 1),
                'Owner': current_owner,
                'Status': '', 'Priority': '', 'Readiness': '', 'Type': '', 'Trust Impact': '', 'Founder Role': '',
                'Overview': '', 'Target': '', 'Risks': '', 'Dependencies': '', 'Next Movement': ''
            }
            continue
            
        if current_project:
            # Parse metadata line
            if "Type:" in text and "Status:" in text:
                parts = text.split(" | ")
                for part in parts:
                    if ":" in part:
                        k, v = part.split(":", 1)
                        k = k.strip()
                        v = v.strip()
                        if k == "Type": current_project["Type"] = v
                        elif k == "Status": current_project["Status"] = v
                        elif k == "Priority": current_project["Priority"] = v
                        elif k == "Trust impact": current_project["Trust Impact"] = v
                        elif k == "Founder role": current_project["Founder Role"] = v
                        elif k == "Readiness": current_project["Readiness"] = v
                continue
                
            # Collect text under headers
            if text.startswith("1. Project overview"):
                current_project['_current_section'] = 'Overview'
            elif text.startswith("3. Target and success"):
                current_project['_current_section'] = 'Target'
            elif text.startswith("4. Clarity map"):
                current_project['_current_section'] = 'Risks'
            elif text.startswith("5. Dependencies"):
                current_project['_current_section'] = 'Dependencies'
            elif text.startswith("6. Next movement"):
                current_project['_current_section'] = 'Next Movement'
            elif text.startswith("2. Stakeholders"):
                current_project['_current_section'] = None
            else:
                section = current_project.get('_current_section')
                if section:
                    current_project[section] += text + "\n"
                    
    if current_project:
        projects.append(current_project)
        
    # Clean up fields
    for p in projects:
        for k in ['Overview', 'Target', 'Risks', 'Dependencies', 'Next Movement']:
            if k in p:
                p[k] = p[k].strip()
        p.pop('_current_section', None)
        
    return projects

def create_excel(projects, output_file):
    wb = openpyxl.Workbook()
    
    # --- COMPANY MAP ---
    ws_map = wb.active
    ws_map.title = "COMPANY MAP"
    
    headers = [
        "ID / Ref", "Item / Project Name", "Owner", "Status", "Priority", 
        "Attention / Readiness", "Next Movement", "Last Reviewed Date", 
        "Type", "Trust Impact", "Founder Role", "Project Overview", 
        "Target & Success", "Known Risks / Open Qs", "Dependencies"
    ]
    
    ws_map.append(headers)
    
    # Styling headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F497D")
    
    for cell in ws_map[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for p in projects:
        row = [
            p['ID'], p['Item'], p['Owner'], p['Status'], p['Priority'],
            p['Readiness'], p['Next Movement'], "",
            p['Type'], p['Trust Impact'], p['Founder Role'], p['Overview'],
            p['Target'], p['Risks'], p['Dependencies']
        ]
        ws_map.append(row)
        
    for row in ws_map.iter_rows(min_row=2, max_col=15, max_row=len(projects)+1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
    # Set column widths
    widths = [8, 30, 10, 15, 10, 15, 30, 15, 12, 12, 25, 40, 40, 40, 30]
    for i, w in enumerate(widths, 1):
        ws_map.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # --- PORTFOLIOS ---
    ws_port = wb.create_sheet("PORTFOLIOS")
    ws_port['A1'] = "Select Owner:"
    ws_port['A1'].font = Font(bold=True)
    ws_port['B1'] = "Kimia" # Default
    
    ws_port['A3'] = "Instructions:"
    ws_port['A4'] = "Since this is an Excel export, filtering logic differs from Google Sheets."
    ws_port['A5'] = "When you import this file into Google Sheets, please replace row 7 with the following formula in A7:"
    ws_port['A6'] = "=FILTER('COMPANY MAP'!A:O, 'COMPANY MAP'!C:C = B1)"
    ws_port['A6'].font = Font(bold=True, color="0000FF")

    # --- CALIBRATION LOG ---
    ws_log = wb.create_sheet("CALIBRATION LOG")
    log_headers = ["Date", "Item Ref", "Decision Made / Notes", "Action Required"]
    ws_log.append(log_headers)
    for cell in ws_log[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_log.column_dimensions['A'].width = 15
    ws_log.column_dimensions['B'].width = 15
    ws_log.column_dimensions['C'].width = 50
    ws_log.column_dimensions['D'].width = 50

    # --- _CONFIG ---
    ws_config = wb.create_sheet("_CONFIG")
    ws_config.sheet_state = 'hidden'
    ws_config.append(["Status", "Priority", "Owner"])
    ws_config.append(["Not Started", "High", "Kimia"])
    ws_config.append(["In Progress", "Medium", "Sepehr"])
    ws_config.append(["Blocked", "Low", "Goksel"])
    ws_config.append(["Parked", "", "Serena"])
    ws_config.append(["Completed", "", "Aryan"])

    # Data Validation for Status, Priority, Owner
    dv_status = DataValidation(type="list", formula1="='_CONFIG'!$A$2:$A$6", allow_blank=True)
    ws_map.add_data_validation(dv_status)
    dv_status.add("D2:D1000")

    dv_priority = DataValidation(type="list", formula1="='_CONFIG'!$B$2:$B$4", allow_blank=True)
    ws_map.add_data_validation(dv_priority)
    dv_priority.add("E2:E1000")
    
    dv_owner = DataValidation(type="list", formula1="='_CONFIG'!$C$2:$C$6", allow_blank=True)
    ws_map.add_data_validation(dv_owner)
    dv_owner.add("C2:C1000")

    # Validation for Portfolios tab Owner selection
    dv_port_owner = DataValidation(type="list", formula1="='_CONFIG'!$C$2:$C$6", allow_blank=True)
    ws_port.add_data_validation(dv_port_owner)
    dv_port_owner.add("B1")

    wb.save(output_file)

if __name__ == '__main__':
    projects = parse_docx('/Users/aryanaghili/Downloads/GOLIVEA Project Intake Review System AA v2.0.docx')
    print(f"Parsed {len(projects)} projects.")
    create_excel(projects, '/Users/aryanaghili/Desktop/Golivea_Operating_Map_Unified.xlsx')
    print("Excel file created at ~/Desktop/Golivea_Operating_Map_Unified.xlsx")
